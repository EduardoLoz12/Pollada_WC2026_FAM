"""
Export Supabase data → polla_mundialera.xlsx
Run: python python/export_excel.py

Sheets generated:
  1. Tabla General   — leaderboard with points breakdown
  2. Pronósticos     — all predictions per match per participant
  3. Partidos        — WC match results
  4. Grupos          — group standings
  5. Goleadores      — top scorers
"""
import os, sys, json, requests
from datetime import datetime, timezone, timedelta
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

load_dotenv()

SB_URL  = os.environ["SUPABASE_URL"]
SB_KEY  = os.environ["SUPABASE_SERVICE_KEY"]
SB_HDR  = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

POINTS  = {"correct_result": 2, "exact_score_bonus": 3, "knockout_bonus": 3,
           "champion": 10, "runner_up": 5, "top_scorer": 5}

OUT_PATH = Path(__file__).parent.parent / "excel" / "polla_mundialera.xlsx"

STAGE_LABEL = {
    "GROUP_STAGE":    "Fase de Grupos",
    "ROUND_OF_32":    "Ronda de 32",
    "ROUND_OF_16":    "Octavos de Final",
    "QUARTER_FINALS": "Cuartos de Final",
    "SEMI_FINALS":    "Semifinales",
    "THIRD_PLACE":    "Tercer Puesto",
    "FINAL":          "Final",
}

# ─── Colors ──────────────────────────────────────────────────────────────────
DARK_BG  = "0A1628"
GOLD     = "F5A623"
GOLD2    = "FFD166"
GREEN    = "27AE60"
RED      = "E74C3C"
WHITE    = "E8F0FE"
SURFACE  = "162845"
MUTED    = "4A5580"


def sb_get(table, params=None):
    r = requests.get(f"{SB_URL}/rest/v1/{table}", headers=SB_HDR, params=params or {})
    r.raise_for_status()
    return r.json()


def col_time(utc_str):
    if not utc_str:
        return ""
    dt = datetime.fromisoformat(utc_str.replace("Z", "+00:00"))
    col = dt - timedelta(hours=5)  # UTC → Colombia
    return col.strftime("%d/%m %H:%M")


def calc_points(preds, matches):
    total = 0
    correct = 0
    exact = 0
    for p in preds:
        m = next((x for x in matches if x["match_id"] == p["match_id"]), None)
        if not m or m.get("status") != "FINISHED" or m.get("home_score") is None:
            continue
        winner = m.get("winner")
        actual = "H" if winner == "HOME_TEAM" else ("A" if winner == "AWAY_TEAM" else "D")
        if p.get("pred_result") == actual:
            total += POINTS["correct_result"]
            correct += 1
            if p.get("pred_home_score") == m["home_score"] and \
               p.get("pred_away_score") == m["away_score"]:
                total += POINTS["exact_score_bonus"]
                exact += 1
            if m.get("stage") != "GROUP_STAGE":
                total += POINTS["knockout_bonus"]
    return total, correct, exact


def style_header(cell, bg=DARK_BG, fg=GOLD, bold=True, size=11):
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="thin", color=GOLD))


def style_cell(cell, bg=SURFACE, fg=WHITE, bold=False, align="left"):
    cell.font      = Font(color=fg, bold=bold, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")


def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(length + 3, max_w))


# ─── Sheet builders ───────────────────────────────────────────────────────────

def build_leaderboard(wb, participants, predictions, matches, special_bets):
    ws = wb.create_sheet("Tabla General")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    headers = ["#", "Nombre", "Avatar", "Pts Total", "Aciertos", "Marcadores Exactos",
               "Pronósticos Hechos", "Campeón", "Subcampeón", "Goleador"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(1, i), bg=DARK_BG, fg=GOLD)
        ws.cell(1, i).value = h

    ws.row_dimensions[1].height = 28

    board = []
    for p in participants:
        my_preds = [x for x in predictions if x["participant_id"] == p["id"]]
        sp = next((s for s in special_bets if s["participant_id"] == p["id"]), {})
        pts, correct, exact = calc_points(my_preds, matches)
        board.append((pts, correct, exact, p, my_preds, sp))

    board.sort(key=lambda x: (-x[0], -x[1]))

    for rank, (pts, correct, exact, p, my_preds, sp) in enumerate(board, 1):
        row = [rank, p["name"], p.get("avatar", "⚽"),
               pts, correct, exact, len(my_preds),
               sp.get("champion",""), sp.get("runner_up",""), sp.get("top_scorer","")]
        ws.append(row)
        r = ws.max_row
        bg = SURFACE if rank % 2 == 0 else "0F1E35"
        for c in range(1, len(row) + 1):
            cell = ws.cell(r, c)
            style_cell(cell, bg=bg, align="center" if c != 2 else "left")
        # Points column gold
        pts_cell = ws.cell(r, 4)
        pts_cell.font = Font(bold=True, color=GOLD2, size=12)
        # Rank medal
        medals = {1: "🥇", 2: "🥈", 3: "🥉"}
        ws.cell(r, 1).value = medals.get(rank, str(rank))

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_predictions(wb, participants, predictions, matches):
    ws = wb.create_sheet("Pronósticos")
    ws.sheet_view.showGridLines = False

    finished = [m for m in matches if m.get("status") == "FINISHED"]
    pending  = [m for m in matches if m.get("status") != "FINISHED"]
    all_m    = finished + pending

    # Header row: match info + participant names
    headers = ["Grupo/Fase", "Partido", "Fecha COL", "Resultado Real"] + \
              [p["name"] for p in participants]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(1, i), size=10)
        ws.cell(1, i).value = h

    ws.row_dimensions[1].height = 25

    for m in all_m:
        stage = STAGE_LABEL.get(m.get("stage",""), m.get("stage",""))
        match_str = f"{m['home_team']} vs {m['away_team']}"
        date_str  = col_time(m.get("kickoff_utc"))
        result    = ""
        if m.get("status") == "FINISHED":
            winner = m.get("winner","")
            w = "LOCAL" if winner=="HOME_TEAM" else ("VISITANTE" if winner=="AWAY_TEAM" else "EMPATE")
            result = f"{m['home_score']}-{m['away_score']} ({w})"

        row_vals = [stage, match_str, date_str, result]

        # Each participant's prediction
        for p in participants:
            pred = next((x for x in predictions
                         if x["participant_id"] == p["id"] and x["match_id"] == m["match_id"]), None)
            if not pred:
                row_vals.append("-")
            else:
                r = pred.get("pred_result","")
                if pred.get("pred_home_score") is not None:
                    row_vals.append(f"{pred['pred_home_score']}-{pred['pred_away_score']} ({r})")
                else:
                    row_vals.append(r)

        ws.append(row_vals)
        r_idx = ws.max_row
        bg = SURFACE if r_idx % 2 == 0 else "0F1E35"
        for c in range(1, len(row_vals) + 1):
            style_cell(ws.cell(r_idx, c), bg=bg, align="center" if c != 2 else "left")

    auto_width(ws)
    ws.freeze_panes = "E2"


def build_matches(wb, matches):
    ws = wb.create_sheet("Partidos")
    ws.sheet_view.showGridLines = False

    headers = ["Fase","Grupo","Fecha COL","Local","Marcador","Visitante","Estado","Ganador"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(1, i))
        ws.cell(1, i).value = h

    for m in matches:
        stage = STAGE_LABEL.get(m.get("stage",""), m.get("stage",""))
        score = f"{m['home_score']}-{m['away_score']}" if m.get("status") == "FINISHED" else "- vs -"
        winner = {"HOME_TEAM": m["home_team"], "AWAY_TEAM": m["away_team"], "DRAW": "Empate"}.get(m.get("winner",""), "")
        ws.append([
            stage, m.get("group_name",""),
            col_time(m.get("kickoff_utc")),
            m["home_team"], score, m["away_team"],
            m.get("status",""), winner,
        ])
        r = ws.max_row
        bg = SURFACE if r % 2 == 0 else "0F1E35"
        for c in range(1, 9):
            style_cell(ws.cell(r, c), bg=bg, align="center" if c != 4 else "right")
        if m.get("status") == "FINISHED":
            ws.cell(r, 5).font = Font(bold=True, color=GREEN, size=11)

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_groups(wb, standings):
    ws = wb.create_sheet("Grupos")
    ws.sheet_view.showGridLines = False

    headers = ["Grupo","Pos","Equipo","PJ","G","E","P","GF","GC","DG","Pts"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(1, i))
        ws.cell(1, i).value = h

    for s in sorted(standings, key=lambda x: (x.get("group_name",""), x.get("position",99))):
        ws.append([
            s.get("group_name","").replace("GROUP_","Grupo "),
            s.get("position",""),
            s.get("team",""),
            s.get("played",0), s.get("won",0), s.get("drawn",0), s.get("lost",0),
            s.get("goals_for",0), s.get("goals_against",0),
            s.get("goal_diff",0), s.get("points",0),
        ])
        r = ws.max_row
        pos = s.get("position", 5)
        bg = SURFACE if r % 2 == 0 else "0F1E35"
        for c in range(1, 12):
            style_cell(ws.cell(r, c), bg=bg, align="center" if c != 3 else "left")
        fg = GREEN if pos <= 2 else (GOLD if pos == 3 else (RED if pos == 4 else WHITE))
        ws.cell(r, 3).font = Font(color=fg, size=10)
        ws.cell(r, 11).font = Font(bold=True, color=GOLD2, size=11)

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_scorers(wb, scorers):
    ws = wb.create_sheet("Goleadores")
    ws.sheet_view.showGridLines = False

    headers = ["#","Jugador","Selección","Goles","Asistencias"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(1, i))
        ws.cell(1, i).value = h

    for i, s in enumerate(sorted(scorers, key=lambda x: -x.get("goals",0)), 1):
        ws.append([i, s.get("player_name",""), s.get("team",""), s.get("goals",0), s.get("assists",0)])
        r = ws.max_row
        bg = SURFACE if r % 2 == 0 else "0F1E35"
        for c in range(1, 6):
            style_cell(ws.cell(r, c), bg=bg, align="center" if c != 2 else "left")
        ws.cell(r, 4).font = Font(bold=True, color=GOLD2, size=12)

    auto_width(ws)
    ws.freeze_panes = "A2"


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Fetching data from Supabase...")
    participants = sb_get("participants", {"order": "created_at"})
    predictions  = sb_get("predictions")
    special_bets = sb_get("special_bets")
    matches      = sb_get("wc_matches", {"order": "kickoff_utc"})
    standings    = sb_get("group_standings", {"order": "group_name,points.desc"})
    scorers      = sb_get("scorers", {"order": "goals.desc", "limit": "50"})

    print(f"  {len(participants)} participantes | {len(predictions)} pronósticos | {len(matches)} partidos")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_leaderboard(wb, participants, predictions, matches, special_bets)
    build_predictions(wb, participants, predictions, matches)
    build_matches(wb, matches)
    build_groups(wb, standings)
    build_scorers(wb, scorers)

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n✅ Guardado en: {OUT_PATH}")
    print(f"   Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")


if __name__ == "__main__":
    main()
